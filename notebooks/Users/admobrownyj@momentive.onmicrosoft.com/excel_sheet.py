# Databricks notebook source
import pandas as pd
import re
import numpy as np
import glob

# COMMAND ----------

def SQL_connection():
    import pyodbc
    import configparser
    import traceback

    config = configparser.ConfigParser()
      #This configuration path should be configured in Blob storage
    config.read("/dbfs/mnt/momentive-configuration/config-file.ini")
    
    server = config.get('sql_db', 'server')
    database = 'CLD-IT-DEV-PIH-DB1'
    username = config.get('sql_db', 'username')
    password = config.get('sql_db', 'password')
  

    driver= "{ODBC Driver 17 for SQL Server}"
    connection_string = 'DRIVER=' + driver + \
              ';SERVER=' + server + \
              ';PORT=1433' + \
              ';DATABASE=' + database + \
              ';UID=' + username + \
              ';PWD=' + password
    
    try:
       sql_conn = pyodbc.connect(connection_string)
       return sql_conn
      # execute query and save data in pandas df
    except Exception as error:
          print("    \u2717 error message: {}".format(error))
      # I found that traceback prints much more detailed error message
          traceback.print_exc()
sql_cursor = SQL_connection()
cursor = sql_cursor.cursor()

# COMMAND ----------

import openpyxl
import os
import csv
import pandas as pd

query = "select blob_src_base_dir, absolute_path from momentive.external_source_folder_structure where blob_base_dir = 'staging' and is_active_folder = 1 and blob_file_type_folder='excel';"
data = pd.read_sql(query, sql_cursor)
data_values = data.values.tolist()

query2 = 'select * from momentive.external_excel_source'
data2 = pd.read_sql(query2, sql_cursor)

# COMMAND ----------

import pandas as pd
file_format_query = "select file_format from momentive.external_source_file_formats where is_active='1'"
file_format_data = pd.read_sql(file_format_query, sql_cursor)

# COMMAND ----------

source_type = 'select source_type from momentive.external_excel_source'
source_type1 = pd.read_sql(source_type, sql_cursor)
list_wel = list(pd.unique(source_type1.source_type))

# COMMAND ----------

query3 = 'select cas_no from [momentive].[product_inscope]'
product_inscope = pd.read_sql(query3, sql_cursor)
CAS_list = list(product_inscope.cas_no)
query4 = 'select * from [momentive].[product_inscope]'
product_list = pd.read_sql(query4, sql_cursor)
nam_prod_list = list(product_list.nam_prod)
bdt_list = list(product_list.bdt)

# COMMAND ----------

import csv
import glob
def write2csv(path, sheet):
  try:
#    logger.info("Executing excel2csv function")
    wb = openpyxl.load_workbook(path)
    sh = wb[sheet]
    head, tail = os.path.split(path)
    filename = path.split('/')[-1].split('.')[0]
    file = head + '/' + 'temp/csv/' + sheet + '.csv'
#    dbutils.fs.rm((absolute_path +'temp/csv/').replace("/dbfs",""),True)
    dbutils.fs.mkdirs((absolute_path +'temp/csv/').replace("/dbfs","dbfs:")) 
    with open(file, 'w', encoding="utf-8") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    return file
  except Exception as e:
    print('Error in write2csv', e)

# COMMAND ----------

def consolidation_output(df1, left_primary, right_primary):
  try:
    final_data = df1.merge(product_list, how='inner', left_on=left_primary, right_on=right_primary)
    final_data.drop(axis=1, labels=[right_primary], inplace=True)
    final_data['spec_id'] = final_data['spec_id'].apply(lambda x: '{0:0>12}'.format(x))
    return final_data
  except Exception as e:
    print('Error in consolidation output', e)

# COMMAND ----------

query_value = "insert into momentive.unstructure_processed_data (product_type,product,data_extract,is_relevant, created,updated) values ('{}','{}','{}','{}',{},{})"

# COMMAND ----------

import json
def key_data_extract_external_source():
  global key_value_df_master_data
  json_list = []
  try:
    if os.path.exists(valid_path + 'relevant_data_files/'):
      files = glob.glob(valid_path + 'relevant_data_files/' + '*.csv')
      for file in files:
        non_rel_data = pd.read_csv(file, encoding='iso-8859-1')
        product = 'Product'
        temp_data = non_rel_data.copy()
        temp_data.drop([product, 'Product_category', 'Component', 'is_relevant'], axis=1, inplace=True)
        df_dict = temp_data.to_json(orient='records', lines=False, force_ascii=False)
        d = json.loads(df_dict)
        for i in range(len(d)):
          b = json.dumps(d[i])
          json_list.append(b)
        key_value_df = pd.DataFrame(json_list, columns =['values']) 
        key_value_df_master = non_rel_data.join(key_value_df)
        key_value_df_master_data = key_value_df_master.loc[:, ['Product_category', product, 'values', 'is_relevant']]
        
  except Exception as e:
    print(e)

# COMMAND ----------

key_data_extract_external_source()

# COMMAND ----------

key_value_df_master_data.head(2)

# COMMAND ----------

def update_operation(query,sql_cursor,cursor):
  cursor.execute(query)
  sql_cursor.commit()


# COMMAND ----------

def unstructure_processed_data(query_value,key_value_df_master_data,sql_cursor,cursor):
  query_value = "insert into momentive.unstructure_processed_data (product_type,product,data_extract,is_relevant, created,updated) values ('{}','{}','{}','{}',{},{})"
  for i in range(key_value_df_master_data.shape[0]):
    insert_query = query_value.format(key_value_df_master_data['Product_category'][i],\
    key_value_df_master_data['Product'][i],key_value_df_master_data['values'][i],  \
    key_value_df_master_data['is_relevant'][i], 'GETDATE()', 'GETDATE()')
#    update_operation(insert_query,sql_cursor,cursor)


# COMMAND ----------

key_value_df_master_data.head(2)

# COMMAND ----------

unstructure_processed_data(query_value,key_value_df_master_data,sql_cursor,cursor)

# COMMAND ----------

def excel_full_delta_load(relevant_data):
  global data_delta
  data_delta = pd.DataFrame()
  try:
    if not os.path.exists(valid_path + 'valid/'):
      dbutils.fs.mkdirs((valid_path +'valid/').replace("/dbfs","dbfs:")) 
      relevant_data.to_csv(valid_path + 'valid/' + 'valid_data.csv', index=None, header=True, encoding='iso-8859-1')
      match_data = pd.DataFrame()
      flag=1
    else:
      mat = glob.glob(valid_path + 'valid/' + '*.csv')
      for m in mat:
        match_data = pd.read_csv(m, encoding='iso-8859-1')
        flag=0
  #        match_data.to_csv(valid_path + 'valid/' + 'valid_data1.csv')
    match_data.replace({r'[^\x00-\x7F]+':''}, regex=True, inplace=True)
#    print('relevant_data count', relevant_data.shape[0])
#    print('match_data_count', match_data.shape[0])
    data_delta = relevant_data.append(match_data).drop_duplicates(keep=False, inplace=True).reset_index(drop=True, inplace=True)
  #    data_delta.reset_index(drop=True, inplace=True)
#    data_delta.drop_duplicates(keep=False, inplace=True)
#    data_delta.reset_index(drop=True, inplace=True)
#    print('data_delta', data_delta.shape[0])
    data_delta1 = data_delta.append(match_data)
    dup = data_delta1.duplicated(keep='first')
    data_delta2 = data_delta1[dup].reset_index(drop=True, inplace=True)
#    data_delta2.reset_index(drop=True, inplace=True)
#    print('data_delta2', data_delta2.shape[0])

    if not data_delta.shape[0]==0 and not flag==1:
      data_to_valid = pd.read_csv(valid_path + 'valid/' + 'valid_data.csv', encoding='iso-8859-1')
      data_to_m = data_delta.append(data_delta2).drop_duplicates(keep=False, inplace=True).reset_index(drop=True, inplace=True)
#      data_to_m.drop_duplicates(keep=False, inplace=True)
#      data_to_m.reset_index(drop=True, inplace=True)
      data_to_v = data_to_valid.append(data_to_m).reset_index(drop=True, inplace=True)
#      data_to_v.reset_index(drop=True, inplace=True)
      dbutils.fs.rm((valid_path +'valid/').replace("/dbfs",""),True)
      data_to_v.to_csv(valid_path + 'valid/' + 'valid_data.csv', index=None, encoding='iso-8859-1')
      data_delta = data_to_m.copy()
      print('data delta after deletion', data_delta.shape[0])
   
    if not data_delta2.shape[0]==0 and not flag==1:
      data_to_v = data_to_valid.append(data_delta2).drop_duplicates(keep=False, inplace=True).reset_index(drop=True, inplace=True)
#      data_to_v.drop_duplicates(keep=False, inplace=True)
#      data_to_v.reset_index(drop=True, inplace=True)
      dbutils.fs.rm((valid_path +'valid/').replace("/dbfs",""),True)
      data_to_v.to_csv(valid_path + 'valid/' + 'valid_data.csv', index=None, encoding='iso-8859-1')
    return data_delta
  except Exception as e:
    print('Error in loading', e)

# COMMAND ----------

def reading_excel_data_from_source(valid_path, files, component_data, primary_column, comp):
  global relevant_data
  regex1 = re.compile(r'(\d+-\d+-\d+)', re.I) #CAS number formatting
  regex2 = re.compile(r'(Y-\d+)', re.I)  #Y-Number formatting
  regex3 = re.compile(r'(\s+/+\s+)', re.I) #
  
  try:
    component_columns = list(component_data['column_name'])
    data_valid_extract = pd.read_csv(files, encoding='iso-8859-1', header=None)
    if data_valid_extract:
      data_valid_extract = data_valid_extract.dropna(how='all',axis=0)
      data_valid_extract.reset_index(drop=True, inplace=True)
      for i in range(data_valid_extract.shape[0]):
        row_list = list(data_valid_extract.loc[i,:])
        start_row_count = list(set(row_list) & set(component_columns))
        if len(start_row_count) >3:
           value_of_column = i
      valid_data = data_valid_extract[int(value_of_column):]
      valid_data = valid_data.rename(columns=valid_data.iloc[0])
      valid_data.drop(valid_data.index[0], inplace=True).reset_index(drop=True, inplace=True)
#      valid_data.reset_index(drop=True, inplace=True)
      valid_data.columns = valid_data.columns.str.replace('\n',' ')
      valid_data.columns = valid_data.columns.str.strip()
      valid_data.columns = valid_data.columns.str.replace(r'[^\x00-\x7F]+', '')
      relevant_data = valid_data.loc[:, component_columns]
      relevant_data.replace({r'[^\x00-\x7F]+':''}, regex=True, inplace=True).drop_duplicates(keep='first', inplace=True).reset_index(drop=True, inplace=True)
#      relevant_data.drop_duplicates(keep='first', inplace=True)
#      relevant_data.reset_index(drop=True, inplace=True)
#      data_delta = excel_full_delta_load(relevant_data)
    return relevant_data
  #    data_delta.to_csv(valid_path + 'valid/' + 'valid_data2.csv', index=None)

  except Exception as e:
    print('Error in reading consolidation', e)

# COMMAND ----------

def data_validation_to_relevant_non_relevant_split(data_delta):
  try:
    query_product_list = 'select * from [momentive].[product_inscope]'
    product_inscope = pd.read_sql(query_product_list, sql_cursor)
    CAS_list = list(product_inscope.cas_no)
  #  query4 = 'select * from [momentive].[product_inscope]'
  #  product_list = pd.read_sql(query4, sql_cursor)
    nam_prod_list = list(product_inscope.nam_prod)
    bdt_list = list(product_inscope.bdt)

    reg_ex = [] 
    reg_ex1 = []
    dbutils.fs.rm((valid_path +'relevant_data_files/').replace("/dbfs",""),True)
    if not data_delta.shape[0]==0:
      for i in range(data_delta.shape[0]):
        product = data_delta.loc[i, primary_column]
        reg_ex = regex1.findall(str(product)) or regex2.findall(str(product)) or regex3.findall(str(product))
        if len(reg_ex)>0:
          data_delta.loc[i, primary_column]=reg_ex[0]
        reg_ex1 = regex1.findall(str(product)) and regex2.findall(str(product))
        if len(reg_ex1)>0:
          data_copy = data_delta.loc[i,:]
          data_delta =  data_delta.append(data_copy, ignore_index=True)
          data_delta.loc[i, primary_column]=reg_ex1[0]
        master_relevant = data_delta.copy()
        master_relevant.rename(columns = {primary_column:'Product'}, inplace=True)
        cas_df = data_delta[primary_column].isin(CAS_list)
        cas_final = data_delta[cas_df]
        cas_final['Product_category'] = 'CAS'
        nam_prod_list_df = data_delta[primary_column].isin(nam_prod_list)
        nam_prod_final = data_delta[nam_prod_list_df]
        nam_prod_final['Product_category'] = 'NAM_PROD'
        bdt_df = data_delta[primary_column].isin(bdt_list)
        bdt_final = data_delta[bdt_df]
        bdt_final['Product_category'] = 'BDT'
        consol_data = pd.concat([cas_final, nam_prod_final, bdt_final])
        consol_data.rename(columns = {primary_column:'Product'}, inplace=True)
        consol_data['Component'] = comp
        consol_data['is_relevant'] = 1
        master_consol_data = consol_data.copy()
        master_consol_data.drop(columns={'Product_category', 'Component', 'is_relevant'}, inplace=True)
        dbutils.fs.mkdirs((valid_path +'relevant_data_files/').replace("/dbfs","dbfs:")) 
        if not consol_data.shape[0]==0:
          relevant_files = consol_data.to_csv(valid_path + 'relevant_data_files/' + 'relevant_data' +'.csv', index=None, header=True)
        final = master_relevant.append(master_consol_data).drop_duplicates(keep=False, inplace=True)
  #          final.drop_duplicates(keep=False, inplace=True)
        final['Component'] = comp
        final['is_relevant'] = 0
        final['Product_category'] = np.nan
        if not final.shape[0]==0:
          final.to_csv(valid_path + 'relevant_data_files/' + 'non_relevant_data' +'.csv', index=None, header=True)   
  except Exception as e:
    print('Error in exception', e)

# COMMAND ----------

def reading_excel_sources(source_type):
  try:
    excel_momentive_source = 'select * from momentive.external_excel_source'
    data_excel_external_source = pd.read_sql(dexcel_momentive_source, sql_cursor)
    dataframe_excel_sources = data_excel_external_source[(data_excel_external_source['source_type']==source_type) & \
                                                         (data_excel_external_source['is_active_folder']==1) & \
                                                         (data_excel_external_source['is_active_column']=='1') & \
                                                         (data_excel_external_source['is_active_sheet']=='1')]
    primary_field = data_excel_external_source[(data_excel_external_source['source_type']==source_type) & \
                                               (data_excel_external_source['is_active_folder']==1) & \
                                               (data_excel_external_source['is_active_column']=='1') & \
                                               (data_excel_external_source['is_active_sheet']=='1') & \
                                               (data_excel_external_source['is_primary']=='1')]
    primary_col = primary_field.column_name.values
    external_sheet = list(pd.unique(dataframe_excel_sources['sheet_name']))
    return dataframe_excel_sources, external_sheet, primary_col
  except Exception as e:
    print('Error in reading sources', e)

# COMMAND ----------

def excel2csv(path, sheet):
  try:
     # write2csv(path, sheet)
      wb = openpyxl.load_workbook(path)
      sh = wb[sheet]
      head, tail = os.path.split(path)
      filename = path.split('/')[-1].split('.')[0]
      file = head + '/' + 'temp/csv/' + sheet + '.csv'
#    dbutils.fs.rm((absolute_path +'temp/csv/').replace("/dbfs",""),True)
      dbutils.fs.mkdirs((absolute_path +'temp/csv/').replace("/dbfs","dbfs:")) 
      with open(file, 'w', encoding="utf-8") as f:
          c = csv.writer(f)
          for r in sh.rows:
              c.writerow([cell.value for cell in r])
      return file
  except Exception as e:
    print('Error in excel2csv', e)

# COMMAND ----------

def read_csv(path, data_gadsl, gadsl_sheet):
  try:
    for sheet in gadsl_sheet:
      gadsl_columns = list(data_gadsl['column_name'])
      reading_consolidation(path, gadsl_columns)
  except Exception as e:
    print('Error in read_csv', e)

# COMMAND ----------

def excel2txt(absolute_path, abs_path, filename, sheet):
  try:
    data_text = pd.read_csv(abs_path, encoding='cp1252')
    file = absolute_path +'temp/temp_all_text/'+ filename +'_'+ sheet+'.txt'
    dbutils.fs.mkdirs((absolute_path +'temp/temp_all_text/').replace("/dbfs","dbfs:")) 
    data_text.to_csv(file)
    
  except Exception as e:
    print('Error in excel2txt', e)

# COMMAND ----------

import openpyxl
import os
import pandas as pd
staging_path = ['/dbfs/mnt/python/GADSL/']
import glob
for absolute_path in staging_path:
  try:
    if '.xlsx' or '.xls' or '.csv' or '.xlsm' in file_format:
      path = glob.glob(absolute_path + '*.*')
      for abs_path in path:
        head, tail = os.path.split(abs_path)
        file_extn = tail.rsplit('.',1)[-1]
        file_name = tail.rsplit('.',1)[0]
        
        if 'csv':
          dbutils.fs.rm((absolute_path +'temp/temp_all_text/').replace("/dbfs",""),True)
          dbutils.fs.rm((absolute_path +'temp/csv/').replace("/dbfs",""),True)
          sheet =str(1)
          excel2txt(absolute_path, abs_path, file_name, sheet)
          text = glob.glob(absolute_path +'temp/temp_all_text/'+'*.txt')
          dbutils.fs.mkdirs((absolute_path +'all_text/').replace("/dbfs","dbfs:")) 
          text_csv = pd.DataFrame()
          for t in text:
            data = pd.read_csv(t, encoding='utf-8')
            text_csv = text_csv.append(data)
            text_csv.to_csv(absolute_path +'all_text/'+file_name + '.txt')
           
        if 'xlsx' or 'xlsm':
          wb = openpyxl.load_workbook(abs_path) 
          allsheets = list(wb.sheetnames)
          dbutils.fs.rm((absolute_path +'temp/temp_all_text/').replace("/dbfs",""),True)
          dbutils.fs.rm((absolute_path +'temp/csv/').replace("/dbfs",""),True)
          for sheet in allsheets:
            excel2csv(abs_path, sheet)
            temp_path = glob.glob(absolute_path+'temp/csv/'+'*.*')
            dbutils.fs.mkdirs((absolute_path +'csv/'+file_name +'/').replace("/dbfs","dbfs:"))
            dbutils.fs.cp((absolute_path +'temp/csv/').replace("/dbfs","dbfs:"), (absolute_path +'csv/'+file_name+'/').replace("/dbfs","dbfs:"), recurse=True)
            for i in range(len(temp_path)):
              excel2txt(absolute_path, temp_path[i], file_name, sheet)
              text_excel = glob.glob(absolute_path +'temp/temp_all_text/'+'*.txt')
              dbutils.fs.mkdirs((absolute_path +'all_text/').replace("/dbfs","dbfs:")) 

              text1 = pd.DataFrame()
              for t in text_excel:
                data = pd.read_csv(t, encoding='utf-8')
                text1 = text1.append(data)
                text1.to_csv(absolute_path +'all_text/'+file_name + '.txt')

  except Exception as e:
    print('Error in main', e)

# COMMAND ----------

def excel_extract2_key_value_pair(valid_path):
  try:
    external_source_data = 'select source_type from momentive.external_excel_source'
    source_type = pd.read_sql(external_source_data, sql_cursor)
    list_components = list(pd.unique(source_type.external_source_data))
    if not list_components:
      print('list is empty')
    else:
      for comp in list_components:
        if comp:
          component_data, component_sheet, primary_col = reading_excel_sources(comp.strip())
          for sheet in component_sheet:
            if sheet:
              valid_files = glob.glob(valid_path + '*.csv')
              for files in valid_files:
                head, tail = os.path.split(files)
                file_name = tail.rsplit('.',1)[0]
                if file_name.strip()==sheet.strip():
                  for primary in primary_col:
                    reading_excel_data_from_source(valid_path, files, component_data, primary, comp)
            else:
              print('sheet is empty')
        else:
          print('component is empty')
    except Exception as e:
    print('exception', e)

# COMMAND ----------

import re
import numpy as np
valid_path = '/dbfs/mnt/test-pih/python/'
excel_extract2_key_value_pair(valid_path)

# COMMAND ----------

def excel_functions():
  SQL_connection()
  valid_path = '/dbfs/mnt/test-pih/python/'
  excel_extract2_key_value_pair(valid_path)
  data_delta = excel_full_delta_load(relevant_data)
  data_validation_to_relevant_non_relevant_split(data_delta)
  key_data_extract_external_source()
  unstructure_processed_data(query_value,key_value_df_master_data,sql_cursor,cursor)
  update_operation(query,sql_cursor,cursor)
  

# COMMAND ----------

data_to_v.head(3)

# COMMAND ----------

relevant_data.shape[0]

# COMMAND ----------

import pandas as pd
import glob
import os

# COMMAND ----------

